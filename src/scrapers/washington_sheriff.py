"""
Washington County Sheriff's Sales (completed foreclosures) scraper.

Source: Washington County Property Records & Taxpayer Services (PRTS) publishes
a monthly "Report of Sheriff's Sales" as an Excel (.xlsx) file in its public
Archive Center.

    List page:   https://www.washingtoncountymn.gov/Archive.aspx?AMID=43
                 ("PRTS - Foreclosure Data - Monthly Reports")
    File (each): https://www.washingtoncountymn.gov/Archive.aspx?ADID={adid}
                 (one .xlsx per month, e.g. ADID=2814 = "2026 Foreclosure Data April")

License / posture: official Washington County government site. Public foreclosure
sale data under the Minnesota Government Data Practices Act. The data lives in the
PRTS archive, NOT on the Sheriff's Civil-Unit page (which says "no list, see
newspapers" — that page is a misleading dead end; the real per-property data is
the monthly XLS published here). No anti-bot terms identified. GREEN per the
data-source audit. We fetch politely.

=== WHAT THIS IS ===
COMPLETED sheriff's sales, recorded monthly, per-property. Same TYPE of signal as
Hennepin/Anoka/Dakota foreclosure (post-auction → starts the redemption clock).
It is NOT a forward calendar of upcoming sales — Washington genuinely doesn't
publish those. Fits the existing "foreclosure" category exactly.

=== XLS STRUCTURE (verified against the April 2026 file, 2026-06-03) ===
Sheet "Sheet1". Row 1 = section banners, Row 2 = column headers, data from row 3.
Columns (by letter):
    A  PID Unformatted          e.g. "2103020330102"  ← join key to TaxParcel/core.parcels
    B  PID Formatted            e.g. "21.030.20.33.0102"
    C  Document #               sheriff's-sale doc number
    D  Date Recorded            the SALE date (a datetime)
    E  Comments                 the SALE AMOUNT, e.g. "$83,413.76"
    F  Purchaser / Grantee 1    who bought it (e.g. REALTY PROS LLC)
    G  Mortgage/Lien Document   original mortgage/lien doc number
    H  Instrument Code          MTG / AL / IOD (codes)
    I  Instrument Code Desc     MORTGAGE / ASSESSMENT LIEN / Image Only Document
    J  Mortgage Dollar Amount   original mortgage amount
    K  Instrument Date          original mortgage date
    L  Date Recorded (orig)     original mortgage recording date
    M  Grantee 1 / Mortgagee    original lender (e.g. WELLS FARGO BANK NA)
    N  Grantee 2 / Mortgagee
    O  Grantor 1 / Mortgagor    THE FORECLOSED OWNER (e.g. CHAVES LUCYANN)
    P  Grantor 2 / Mortgagor

=== INSTRUMENT-TYPE SPLIT (mirrors the Anoka HOA-vs-bank lesson) ===
The file mixes three record kinds, distinguished by the Instrument Code Desc (col I):
  * MORTGAGE            — a bank foreclosure. Has lender (M) and borrower (O).
  * ASSESSMENT LIEN     — an HOA / association lien foreclosure. The "lender" (M)
                          is a person and the "borrower" (O) is the association,
                          or vice-versa; there is no mortgage dollar amount (J empty).
  * Image Only Document — a bare lien image with essentially no structured party
                          data (no owner, no amount). We KEEP it (it's still a real
                          sheriff's-sale record) but it will be sparse / unmatched.
We do not try to force a single owner convention onto all three; we record the
instrument type and store whatever parties are present.

=== KEY / DEDUP NOTE ===
Two rows can share a Document # (e.g. the April file's rows 8 & 9 both have doc
4504143 but different PIDs — a sale covering two adjoining parcels). So the stable
unique identifier is PID + Document#, NOT the document number alone. We build
source_id = "{pid}-{docnum}" and parcel_id = "WASHINGTON-FC-{pid}".

=== ARCHITECTURE ===
fetch():
  1. GET the archive list page (AMID=43).
  2. Parse every <a href="Archive.aspx?ADID=NNNN"> whose label looks like a
     "YYYY Foreclosure Data Month" entry; capture (adid, year, month, label).
  3. Keep only the most recent N months (settings-driven; default a rolling
     window). Download each ADID .xlsx and parse its sale rows.
parse():  each sale row → DistressEventInsert(event_type="foreclosure").
write():  resolve_parcel + write_events_dedup (mirrors Anoka exactly).

Enrichment (owner / market value / homestead) is a SEPARATE step: a
washington_foreclosure_enrichment job that PID-joins these rows to the Washington
TaxParcel layer loaded in core.parcels — exactly as Dakota/Anoka enrichment works.
This scraper just lands the honest sheriff-sale signal.
"""

from __future__ import annotations

import asyncio
import io
import re
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any, ClassVar

import httpx
from bs4 import BeautifulSoup

from src.config import settings
from src.models.parcel import ParcelUpsert
from src.models.signal import DistressEventInsert
from src.scrapers.base_scraper import BaseScraper
from src.services.event_writer import write_events_dedup
from src.services.parcel_resolver import resolve_parcel
from src.utils.errors import ParseError, SourceUnavailableError
from src.utils.logger import logger


_BASE = "https://www.washingtoncountymn.gov"
_LIST_URL = f"{_BASE}/Archive.aspx?AMID=43"
_FILE_URL = f"{_BASE}/Archive.aspx?ADID={{adid}}"

# How many of the most-recent monthly files to ingest per run. Each file is one
# month of completed sales. A rolling window keeps the run fast and the data
# current; older months are already in the DB from prior runs (write_events_dedup
# is insert-only, so re-reading them would be a no-op anyway). Override via
# settings if a deeper backfill is ever wanted.
_DEFAULT_MONTHS_LIMIT = 6

# Politeness: small delay between file downloads.
_FILE_DELAY_SECONDS = 0.5

# Map full month names to numbers, for sorting the archive entries by date.
_MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}

# Recognizes labels like:
#   "2026 Foreclosure Data April (XLS)"
#   "2022 Foreclosure Data - January"
#   "2024 Foreclosure Data March "
# Captures the 4-digit year and the month name; tolerates an optional dash,
# trailing "(XLS)", and extra whitespace.
_RE_LABEL = re.compile(
    r"(20\d{2})\s+Foreclosure\s+Data\s*-?\s*([A-Za-z]+)",
    re.I,
)
_RE_ADID = re.compile(r"ADID=(\d+)", re.I)

_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)
_BROWSER_HEADERS = {
    "User-Agent": _USER_AGENT,
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/avif,image/webp,*/*;q=0.8"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}


def _safe_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _safe_decimal(value: Any) -> Decimal | None:
    """Parse a money value that may arrive as a number or a "$83,413.76" string."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            d = Decimal(str(value))
            return d if d >= 0 else None
        except (InvalidOperation, ValueError):
            return None
    s = str(value).strip()
    if not s:
        return None
    # Strip $, commas, and any stray non-numeric prose; keep digits and one dot.
    cleaned = re.sub(r"[^0-9.]", "", s.replace(",", ""))
    if not cleaned or cleaned == ".":
        return None
    try:
        d = Decimal(cleaned)
        return d if d >= 0 else None
    except (InvalidOperation, ValueError):
        return None


def _cell_date(value: Any) -> date | None:
    """Col D is normally a datetime from openpyxl; tolerate strings too."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _norm_pid(raw: Any) -> str | None:
    """Normalize the unformatted PID (col A): strip, drop internal whitespace.

    Washington's unformatted PID is a plain numeric string (e.g. 2103020330102).
    We keep it verbatim (minus whitespace) so it joins cleanly to the TaxParcel
    PIN. openpyxl may hand us an int for an all-digits cell, so coerce to str.
    """
    if raw is None:
        return None
    if isinstance(raw, float) and raw.is_integer():
        raw = int(raw)
    s = "".join(str(raw).split())
    return s or None


class WashingtonSheriffScraper(BaseScraper[dict[str, Any], DistressEventInsert]):
    """Washington County completed sheriff's sales — monthly PRTS Excel files."""

    source_name: ClassVar[str] = "washington_sheriff"
    signal_type: ClassVar[str] = "foreclosure"
    county_code: ClassVar[str] = "washington"

    # ---- Fetch: list page → monthly XLS downloads → row dicts ----

    async def fetch(self, trigger: str) -> list[dict[str, Any]]:
        months_limit = getattr(
            settings, "washington_sheriff_months_limit", _DEFAULT_MONTHS_LIMIT
        ) or _DEFAULT_MONTHS_LIMIT

        timeout = httpx.Timeout(connect=20.0, read=90.0, write=30.0, pool=30.0)
        all_rows: list[dict[str, Any]] = []

        async with httpx.AsyncClient(
            timeout=timeout,
            headers=_BROWSER_HEADERS,
            follow_redirects=True,
        ) as client:
            # 1. GET the archive list page (with a couple of retries).
            resp = None
            last_err: Exception | None = None
            for attempt in range(3):
                try:
                    resp = await client.get(_LIST_URL)
                    break
                except httpx.HTTPError as e:
                    last_err = e
                    logger.warning(
                        "Washington list GET attempt failed",
                        source=self.source_name,
                        attempt=attempt + 1,
                        error_type=type(e).__name__,
                    )
                    await asyncio.sleep(2.0)
            if resp is None:
                raise SourceUnavailableError(
                    f"Washington archive list GET failed after retries: "
                    f"{type(last_err).__name__}: {last_err!r}",
                    source=self.source_name,
                )
            if resp.status_code != 200:
                raise SourceUnavailableError(
                    f"Washington archive list returned {resp.status_code}",
                    source=self.source_name,
                )

            # 2. Parse the archive entries: (adid, year, month, label).
            entries = self._parse_archive_list(resp.text)
            if not entries:
                raise ParseError(
                    "No 'YYYY Foreclosure Data Month' entries found on the "
                    "Washington archive list page.",
                    source=self.source_name,
                )

            # Most recent first, then keep only the rolling window.
            entries.sort(key=lambda e: (e["year"], e["month"]), reverse=True)
            selected = entries[:months_limit]
            logger.info(
                "Washington archive parsed",
                source=self.source_name,
                total_files=len(entries),
                selected=len(selected),
                newest=f"{selected[0]['year']}-{selected[0]['month']:02d}"
                if selected else None,
            )

            # 3. Download + parse each selected monthly XLS.
            for entry in selected:
                url = _FILE_URL.format(adid=entry["adid"])
                try:
                    await asyncio.sleep(_FILE_DELAY_SECONDS)
                    file_resp = await client.get(url)
                except httpx.HTTPError as e:
                    logger.warning(
                        "Washington file download failed; skipping month",
                        source=self.source_name,
                        adid=entry["adid"],
                        label=entry["label"],
                        error_type=type(e).__name__,
                    )
                    continue
                if file_resp.status_code != 200:
                    logger.warning(
                        "Washington file non-200; skipping month",
                        source=self.source_name,
                        adid=entry["adid"],
                        status_code=file_resp.status_code,
                    )
                    continue

                rows = self._parse_xlsx(
                    file_resp.content, entry["year"], entry["month"], entry["label"]
                )
                logger.info(
                    "Washington month parsed",
                    source=self.source_name,
                    label=entry["label"],
                    rows=len(rows),
                )
                all_rows.extend(rows)

        logger.info(
            "Washington fetch complete",
            source=self.source_name,
            total_rows=len(all_rows),
        )
        return all_rows

    # ---- HTML parsing: the archive list ----

    def _parse_archive_list(self, html: str) -> list[dict[str, Any]]:
        """Find every monthly-foreclosure-file link on the AMID=43 page.

        Each entry is an <a href="Archive.aspx?ADID=NNNN"> whose visible text
        is like "2026 Foreclosure Data April (XLS)". We read the link text
        (the <span> inside the anchor) rather than the href to get the label.
        """
        soup = BeautifulSoup(html, "lxml")
        entries: list[dict[str, Any]] = []
        seen_adids: set[str] = set()

        for link in soup.find_all("a", href=_RE_ADID):
            href = link.get("href", "")
            m_adid = _RE_ADID.search(href)
            if not m_adid:
                continue
            adid = m_adid.group(1)

            label = link.get_text(" ", strip=True)
            m_label = _RE_LABEL.search(label)
            if not m_label:
                # Not a "YYYY Foreclosure Data Month" link (e.g. "All Archives").
                continue

            year = int(m_label.group(1))
            month_name = m_label.group(2).lower()
            month = _MONTHS.get(month_name)
            if month is None:
                # Unrecognized month token — skip rather than guess.
                logger.warning(
                    "Washington archive: unrecognized month token",
                    source=self.source_name,
                    label=label,
                    token=month_name,
                )
                continue

            if adid in seen_adids:
                continue
            seen_adids.add(adid)

            entries.append({
                "adid": adid,
                "year": year,
                "month": month,
                "label": label,
            })
        return entries

    # ---- XLSX parsing: one monthly file → sale-row dicts ----

   def _parse_xlsx(
        self, content: bytes, year: int, month: int, label: str
    ) -> list[dict[str, Any]]:
        """Parse a monthly Report-of-Sheriff's-Sales workbook into row dicts.

        Washington's file layout is NOT consistent across months. Two known shapes:
          * Some files put the data on the first sheet ("Sheet1") with the column
            header on row 2 and data from row 3.
          * Others carry an instructions sheet first ("Read for Information on
            Content") and the real data on a SECOND sheet (with a stale tab name
            like "January 31, 2017"), header on row 7, data from row 8.
        Both share the same A-P column layout. So rather than hardcode a sheet
        index and a start row, we DISCOVER the header: scan every sheet for the
        row whose column A begins with "PID", then read data rows after it. This
        is robust to both layouts and to future template tweaks.
        """
        # Local import: openpyxl is only needed by this scraper. Importing at
        # module top would make the whole scrapers package require it.
        import openpyxl

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(content), data_only=True, read_only=True
            )
        except Exception as e:
            logger.warning(
                "Washington XLSX failed to open; skipping month",
                source=self.source_name,
                label=label,
                error_type=type(e).__name__,
                error=str(e)[:200],
            )
            return []

        # Find the sheet + header row by locating column A == "PID...".
        data_ws = None
        header_row = None
        try:
            for ws in wb.worksheets:
                # Only the first ~20 rows can plausibly hold the header banner.
                for r, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=20, max_col=1, values_only=True),
                    start=1,
                ):
                    a = row[0] if row else None
                    if a is not None and str(a).strip().upper().startswith("PID"):
                        data_ws = ws
                        header_row = r
                        break
                if data_ws is not None:
                    break
        except Exception as e:
            logger.warning(
                "Washington XLSX header scan failed; skipping month",
                source=self.source_name,
                label=label,
                error_type=type(e).__name__,
            )
            try:
                wb.close()
            except Exception:
                pass
            return []

        if data_ws is None or header_row is None:
            logger.warning(
                "Washington XLSX: no PID header row found; skipping month",
                source=self.source_name,
                label=label,
                sheets=wb.sheetnames,
            )
            try:
                wb.close()
            except Exception:
                pass
            return []

        rows: list[dict[str, Any]] = []
        # Data starts on the row AFTER the header. Columns A..P = index 0..15.
        for excel_row in data_ws.iter_rows(min_row=header_row + 1, values_only=True):
            # Guard against short/empty rows.
            def col(i: int) -> Any:
                return excel_row[i] if i < len(excel_row) else None

            pid = _norm_pid(col(0))            # A
            if not pid:
                # No parcel id → can't key it; skip (blank trailing rows, etc.).
                continue

            docnum = _safe_str(col(2))         # C
            sale_date = _cell_date(col(3))     # D
            sale_amount = col(4)               # E (money string/number)
            purchaser = _safe_str(col(5))      # F
            instrument_code = _safe_str(col(7))    # H (MTG/AL/IOD/CIC)
            instrument_desc = _safe_str(col(8))    # I (MORTGAGE/ASSESSMENT LIEN/...)
            mortgage_amount = col(9)               # J
            orig_lender = _safe_str(col(12))   # M (Grantee/Mortgagee/Orig Lender)
            owner = _safe_str(col(14))         # O (Grantor/Mortgagor/Borrower)

            rows.append({
                "pid": pid,
                "pid_formatted": _safe_str(col(1)),   # B
                "docnum": docnum,
                "sale_date": sale_date,
                "sale_amount": sale_amount,
                "purchaser": purchaser,
                "instrument_code": instrument_code,
                "instrument_desc": instrument_desc,
                "mortgage_amount": mortgage_amount,
                "orig_lender": orig_lender,
                "owner": owner,
                "_file_year": year,
                "_file_month": month,
                "_file_label": label,
            })
        try:
            wb.close()
        except Exception:
            pass
        return rows

    # ---- Parse rows → signals ----

    async def parse(
        self, raw_records: list[dict[str, Any]]
    ) -> list[DistressEventInsert]:
        signals: list[DistressEventInsert] = []

        for r in raw_records:
            pid = r.get("pid")
            if not pid:
                continue
            sale_date = r.get("sale_date")
            if sale_date is None:
                # No sale date → can't form a foreclosure event. Skip honestly.
                continue

            docnum = r.get("docnum") or "nodoc"
            parcel_id = f"WASHINGTON-FC-{pid}"
            # PID + Document# is the stable unique key (two parcels can share a doc).
            source_id = f"{pid}-{docnum}"

            instrument_desc = (r.get("instrument_desc") or "").strip()
            instrument_upper = instrument_desc.upper()
            is_mortgage = "MORTGAGE" in instrument_upper
            is_assessment = "ASSESSMENT" in instrument_upper

            # Subtype mirrors the bank-vs-HOA split (the Anoka lesson).
            if is_mortgage:
                event_subtype = "mortgage_foreclosure"
            elif is_assessment:
                event_subtype = "assessment_lien"
            else:
                # Image Only Document or anything else: keep, label generically.
                event_subtype = "other_lien"

            sale_amount = _safe_decimal(r.get("sale_amount"))

            # Severity: completed sales are post-auction (redemption window).
            # Mortgage foreclosures are the core investor signal → medium;
            # assessment liens and image-only records → low.
            severity = "medium" if is_mortgage else "low"

            owner = _safe_str(r.get("owner"))
            purchaser = _safe_str(r.get("purchaser"))
            orig_lender = _safe_str(r.get("orig_lender"))

            title_bits = ["Completed sheriff foreclosure sale"]
            if owner:
                title_bits.append(f"— {owner}")
            title = " ".join(title_bits)[:500]

            desc_parts = [
                f"Completed Washington County sheriff's sale recorded "
                f"{sale_date.isoformat()}."
            ]
            if instrument_desc:
                desc_parts.append(f"Type: {instrument_desc}.")
            if owner:
                desc_parts.append(f"Mortgagor/owner: {owner}.")
            if purchaser:
                desc_parts.append(f"Purchaser: {purchaser}.")
            if orig_lender:
                desc_parts.append(f"Original lender: {orig_lender}.")
            if sale_amount is not None:
                desc_parts.append(f"Sale amount: ${sale_amount:,.0f}.")
            description = " ".join(desc_parts)[:2000]

            signals.append(DistressEventInsert(
                parcel_id=parcel_id,
                event_type="sheriff_sale",
                event_subtype=event_subtype,
                event_date=sale_date,
                event_value=sale_amount,
                source=self.source_name,
                source_id=source_id,
                severity=severity,  # type: ignore[arg-type]
                title=title,
                description=description,
                raw_data={
                    "sale": {
                        "pid": pid,
                        "pid_formatted": r.get("pid_formatted"),
                        "document_number": r.get("docnum"),
                        "sale_date": sale_date.isoformat(),
                        "sale_amount": str(sale_amount) if sale_amount is not None else None,
                        "purchaser": purchaser,
                        "instrument_code": r.get("instrument_code"),
                        "instrument_desc": instrument_desc or None,
                        "mortgage_amount": (
                            str(_safe_decimal(r.get("mortgage_amount")))
                            if _safe_decimal(r.get("mortgage_amount")) is not None
                            else None
                        ),
                        "original_lender": orig_lender,
                        "owner": owner,
                        "source_file": r.get("_file_label"),
                    },
                    # Enrichment (washington_foreclosure_enrichment) will add a
                    # "detail" object with gis_* fields by PID-joining TaxParcel.
                    "_source": self.source_name,
                },
                observed_at=datetime.now(timezone.utc),
            ))

        return signals

    # ---- Write (mirror Anoka: resolve parcels + dedup events) ----

    async def write(
        self, signals: list[DistressEventInsert]
    ) -> tuple[int, int, int]:
        if not signals:
            return 0, 0, 0

        unique_parcels: dict[str, ParcelUpsert] = {}
        for ev in signals:
            if ev.parcel_id in unique_parcels:
                continue
            raw = ev.raw_data or {}
            sale = raw.get("sale") or {}

            unique_parcels[ev.parcel_id] = ParcelUpsert(
                parcel_id=ev.parcel_id,
                county_code=self.county_code,
                state="MN",
                # The sheriff file carries no situs address; address/owner/value
                # come from the TaxParcel enrichment join (by PID). Leave blank
                # here rather than fabricate.
                address=None,
                city=None,
                zip=None,
                raw_data={"washington_foreclosure": sale, "_source": self.source_name},
                data_sources=[self.source_name],
                last_observed_at=datetime.now(timezone.utc),
            )

        parcels_failed = 0
        for payload in unique_parcels.values():
            if resolve_parcel(payload) is None:
                parcels_failed += 1

        new_events, failed_events = write_events_dedup(signals)
        logger.info(
            "Washington write complete",
            source=self.source_name,
            parcels=len(unique_parcels),
            events_new=new_events,
            failed=failed_events + parcels_failed,
        )
        return new_events, 0, failed_events + parcels_failed


__all__ = ["WashingtonSheriffScraper"]
